from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
from flask_cors import CORS

app = Flask(__name__)
CORS(app, resources={r"/registrar": {"origins": "*", "methods": ["POST"]}})

ARQUIVO_EXCEL = os.path.join(os.path.dirname(os.path.abspath(__file__)), "registros.xlsx")

if not os.path.exists(ARQUIVO_EXCEL):
    print("Arquivo Excel não encontrado, criando um novo...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Visitantes"
    ws.append(["Nome", "Cidade", "Idade", "Data", "Tema"])
    wb.save(ARQUIVO_EXCEL)

@app.route('/registrar', methods=['POST'])
def registrar():
    dados = request.get_json()
    print("Dados recebidos:", dados)  # <-- Adicione esta linha
    nome = dados.get("visitante", "").strip()
    cidade = dados.get("cidade", "").strip()
    idade = dados.get("idade", "").strip()
    tema = dados.get("tema", "Exposição: 60 anos do Instituto").strip()
    data = datetime.now().strftime("%d/%m/%Y %H:%M")

    if not nome or not cidade:
        return jsonify({"status": "erro", "mensagem": "Campos obrigatórios faltando"}), 400

    print('PROCURANDO ARQUIVO')

    try:
        print('Arquivo encontrado: ', ARQUIVO_EXCEL)
        wb = load_workbook(ARQUIVO_EXCEL)
        ws = wb.active
        ws.append([nome, cidade, idade, data, tema])
        wb.save(ARQUIVO_EXCEL)
    except Exception as e:
        print("Erro ao salvar no Excel:", e)
        return jsonify({"status": "erro", "mensagem": "Erro ao salvar registro"}), 500

    return jsonify({"status": "sucesso", "mensagem": "Registro salvo com sucesso!"}), 200
    
if __name__ == '__main__':
    print("Iniciando o servidor Flask...")
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))